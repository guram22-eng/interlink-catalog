import os
from flask import Flask, request, jsonify, redirect, render_template_string
from flask_cors import CORS
import psycopg2
from psycopg2.extras import RealDictCursor
from openpyxl import load_workbook

app = Flask(__name__)
CORS(app)

DB_HOST = os.environ.get("DB_HOST")
DB_NAME = os.environ.get("DB_NAME")
DB_USER = os.environ.get("DB_USER")
DB_PASS = os.environ.get("DB_PASS")
DB_PORT = os.environ.get("DB_PORT", "5432")

ADMIN_LOGIN = os.environ.get("ADMIN_LOGIN", "admin")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "1234")

SERIES_LIST = ["MSZ-LN", "MSZ-EF", "MSZ-AY", "MSZ-HR", "MS-GF"]

MXZ_LIST = [
    "MXZ-2F33VF",
    "MXZ-2F42VF",
    "MXZ-2F53VF",
    "MXZ-3F54VF",
    "MXZ-3F68VF",
    "MXZ-4F72VF",
    "MXZ-4F80VF",
    "MXZ-5F102VF"
]


def get_conn():
    return psycopg2.connect(
        host=DB_HOST,
        database=DB_NAME,
        user=DB_USER,
        password=DB_PASS,
        port=DB_PORT,
        cursor_factory=RealDictCursor
    )


def check_admin():
    login = request.args.get("login") or request.form.get("login")
    password = request.args.get("password") or request.form.get("password")
    return login == ADMIN_LOGIN and password == ADMIN_PASSWORD


def admin_url():
    return f"/admin?login={ADMIN_LOGIN}&password={ADMIN_PASSWORD}"


def to_int(value):
    try:
        return int(value or 0)
    except Exception:
        return 0


def calc_total(category, indoor_price, outdoor_price, mxz_price):
    if category == "multisplit":
        return mxz_price
    return indoor_price + outdoor_price


ADMIN_HTML = """
<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<title>Interlink Catalog Admin</title>
<style>
body{
  font-family:Arial,sans-serif;
  max-width:1500px;
  margin:20px auto;
  background:#f8fafc;
  color:#020617;
}
h1{font-size:36px;margin:0 0 18px;}
h2{font-size:28px;margin:0 0 16px;}
.box{
  background:#fff;
  border:1px solid #e5e7eb;
  border-radius:18px;
  padding:18px;
  margin-bottom:22px;
  box-shadow:0 10px 26px rgba(15,23,42,.06);
}
.grid{
  display:grid;
  grid-template-columns:repeat(4,1fr);
  gap:12px;
}
.field label{
  display:block;
  margin-bottom:5px;
  font-size:14px;
  color:#475569;
}
input,select{
  width:100%;
  height:46px;
  padding:8px 11px;
  border:1px solid #cbd5e1;
  border-radius:8px;
  font-size:16px;
  background:#fff;
}
button{
  padding:11px 16px;
  border:0;
  border-radius:8px;
  font-size:16px;
  font-weight:800;
  cursor:pointer;
}
.add{background:#16a34a;color:#fff;}
.save{background:#2563eb;color:#fff;}
.delete{background:#dc2626;color:#fff;}
.item{
  background:#fff;
  border:1px solid #e5e7eb;
  border-radius:16px;
  padding:14px;
  margin-bottom:14px;
}
.actions{
  display:flex;
  gap:10px;
  margin-top:12px;
}
.total{
  font-size:22px;
  color:#16a34a;
  font-weight:900;
  padding-top:8px;
}
.import-note{
  margin:0 0 10px;
  color:#64748b;
  font-size:14px;
}
.hidden{
  display:none !important;
}
@media(max-width:900px){
  .grid{grid-template-columns:1fr;}
}
</style>
</head>
<body>

<h1>Interlink Catalog Admin</h1>

<div class="box">
  <h2>Добавить товар</h2>

  <form method="POST" action="/admin/add" class="catalog-form">
    <input type="hidden" name="login" value="{{login}}">
    <input type="hidden" name="password" value="{{password}}">

    <div class="grid">
      <div class="field">
        <label>Категория</label>
        <select name="category" class="category-select">
          <option value="split">split</option>
          <option value="multisplit">multisplit</option>
        </select>
      </div>

      <div class="field split-field">
        <label>Серия</label>
        <select name="series">
          <option value="">Выберите серию</option>
          {% for s in series %}
          <option value="{{s}}">{{s}}</option>
          {% endfor %}
        </select>
      </div>

      <div class="field">
        <label>Статус</label>
        <select name="status">
          <option>В наличии</option>
          <option>Под заказ</option>
        </select>
      </div>

      <div class="field multi-field">
        <label>MXZ модель</label>
        <select name="mxz_model">
          <option value="">Выберите MXZ</option>
          {% for m in mxz %}
          <option value="{{m}}">{{m}}</option>
          {% endfor %}
        </select>
      </div>

      <div class="field split-field">
        <label>Внутренний блок</label>
        <input name="indoor" placeholder="MSZ-LN35VG2">
      </div>

      <div class="field split-field">
        <label>Цена внутреннего</label>
        <input name="indoor_price" type="number" value="0">
      </div>

      <div class="field split-field">
        <label>Наружный блок</label>
        <input name="outdoor" placeholder="MUZ-LN35VG">
      </div>

      <div class="field split-field">
        <label>Цена наружного</label>
        <input name="outdoor_price" type="number" value="0">
      </div>

      <div class="field multi-field">
        <label>Цена MXZ</label>
        <input name="mxz_price" type="number" value="0">
      </div>
    </div>

    <br>
    <button class="add">Добавить</button>
  </form>
</div>

<div class="box">
  <h2>Импорт Excel</h2>
  <p class="import-note">
    Excel очищает старый каталог и загружает новый.
    Колонки: category, series, indoor, indoor_price, outdoor, outdoor_price, mxz_model, mxz_price, status
  </p>

  <form method="POST" action="/admin/upload" enctype="multipart/form-data">
    <input type="hidden" name="login" value="{{login}}">
    <input type="hidden" name="password" value="{{password}}">
    <input type="file" name="file" accept=".xlsx" required>
    <br><br>
    <button class="add">Загрузить Excel</button>
  </form>
</div>

<div class="box">
  <h2>Каталог</h2>

  {% for item in items %}
  <div class="item">
    <form method="POST" action="/admin/update/{{item.id}}" class="catalog-form">
      <input type="hidden" name="login" value="{{login}}">
      <input type="hidden" name="password" value="{{password}}">

      <div class="grid">
        <div class="field">
          <label>Категория</label>
          <select name="category" class="category-select">
            <option value="split" {% if item.category == 'split' %}selected{% endif %}>split</option>
            <option value="multisplit" {% if item.category == 'multisplit' %}selected{% endif %}>multisplit</option>
          </select>
        </div>

        <div class="field split-field">
          <label>Серия</label>
          <select name="series">
            <option value="">Выберите серию</option>
            {% for s in series %}
            <option value="{{s}}" {% if item.series == s %}selected{% endif %}>{{s}}</option>
            {% endfor %}
          </select>
        </div>

        <div class="field split-field">
          <label>Внутренний</label>
          <input name="indoor" value="{{item.indoor or ''}}">
        </div>

        <div class="field split-field">
          <label>Цена внутр.</label>
          <input name="indoor_price" type="number" value="{{item.indoor_price or 0}}">
        </div>

        <div class="field split-field">
          <label>Наружный</label>
          <input name="outdoor" value="{{item.outdoor or ''}}">
        </div>

        <div class="field split-field">
          <label>Цена наруж.</label>
          <input name="outdoor_price" type="number" value="{{item.outdoor_price or 0}}">
        </div>

        <div class="field multi-field">
          <label>MXZ модель</label>
          <select name="mxz_model">
            <option value="">Выберите MXZ</option>
            {% for m in mxz %}
            <option value="{{m}}" {% if item.mxz_model == m %}selected{% endif %}>{{m}}</option>
            {% endfor %}
          </select>
        </div>

        <div class="field multi-field">
          <label>Цена MXZ</label>
          <input name="mxz_price" type="number" value="{{item.mxz_price or 0}}">
        </div>

        <div class="field">
          <label>Статус</label>
          <select name="status">
            <option {% if item.status == 'В наличии' %}selected{% endif %}>В наличии</option>
            <option {% if item.status == 'Под заказ' %}selected{% endif %}>Под заказ</option>
          </select>
        </div>

        <div class="field">
          <label>Итого</label>
          <div class="total">{{item.total or 0}} ₾</div>
        </div>
      </div>

      <div class="actions">
        <button class="save">Сохранить</button>
    </form>

        <form method="POST" action="/admin/delete/{{item.id}}" onsubmit="return confirm('Удалить товар?');">
          <input type="hidden" name="login" value="{{login}}">
          <input type="hidden" name="password" value="{{password}}">
          <button class="delete">Удалить</button>
        </form>
      </div>
  </div>
  {% endfor %}
</div>

<script>
function updateFormVisibility(form){
  const category = form.querySelector(".category-select").value;
  const splitFields = form.querySelectorAll(".split-field");
  const multiFields = form.querySelectorAll(".multi-field");

  if(category === "multisplit"){
    splitFields.forEach(el => el.classList.add("hidden"));
    multiFields.forEach(el => el.classList.remove("hidden"));
  }else{
    splitFields.forEach(el => el.classList.remove("hidden"));
    multiFields.forEach(el => el.classList.add("hidden"));
  }
}

document.querySelectorAll(".catalog-form").forEach(form => {
  updateFormVisibility(form);
  form.querySelector(".category-select").addEventListener("change", () => {
    updateFormVisibility(form);
  });
});
</script>

</body>
</html>
"""


@app.route("/")
def home():
    return "Interlink Catalog API работает"


@app.route("/admin")
def admin():
    if not check_admin():
        return "Access denied"

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM catalog ORDER BY id DESC")
    items = cur.fetchall()
    cur.close()
    conn.close()

    return render_template_string(
        ADMIN_HTML,
        items=items,
        login=request.args.get("login"),
        password=request.args.get("password"),
        series=SERIES_LIST,
        mxz=MXZ_LIST
    )


@app.route("/admin/add", methods=["POST"])
def add_item():
    if not check_admin():
        return "Access denied"

    d = request.form
    category = d.get("category", "split")

    if category == "multisplit":
        series = ""
        indoor = ""
        indoor_price = 0
        outdoor = ""
        outdoor_price = 0
        mxz_model = d.get("mxz_model", "")
        mxz_price = to_int(d.get("mxz_price"))
    else:
        series = d.get("series", "")
        indoor = d.get("indoor", "")
        indoor_price = to_int(d.get("indoor_price"))
        outdoor = d.get("outdoor", "")
        outdoor_price = to_int(d.get("outdoor_price"))
        mxz_model = ""
        mxz_price = 0

    total = calc_total(category, indoor_price, outdoor_price, mxz_price)

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO catalog
        (category, series, indoor, indoor_price, outdoor, outdoor_price, total, mxz_model, mxz_price, status)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        category,
        series,
        indoor,
        indoor_price,
        outdoor,
        outdoor_price,
        total,
        mxz_model,
        mxz_price,
        d.get("status", "В наличии")
    ))

    conn.commit()
    cur.close()
    conn.close()

    return redirect(admin_url())


@app.route("/admin/update/<int:item_id>", methods=["POST"])
def update_item(item_id):
    if not check_admin():
        return "Access denied"

    d = request.form
    category = d.get("category", "split")

    if category == "multisplit":
        series = ""
        indoor = ""
        indoor_price = 0
        outdoor = ""
        outdoor_price = 0
        mxz_model = d.get("mxz_model", "")
        mxz_price = to_int(d.get("mxz_price"))
    else:
        series = d.get("series", "")
        indoor = d.get("indoor", "")
        indoor_price = to_int(d.get("indoor_price"))
        outdoor = d.get("outdoor", "")
        outdoor_price = to_int(d.get("outdoor_price"))
        mxz_model = ""
        mxz_price = 0

    total = calc_total(category, indoor_price, outdoor_price, mxz_price)

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        UPDATE catalog
        SET category=%s,
            series=%s,
            indoor=%s,
            indoor_price=%s,
            outdoor=%s,
            outdoor_price=%s,
            total=%s,
            mxz_model=%s,
            mxz_price=%s,
            status=%s
        WHERE id=%s
    """, (
        category,
        series,
        indoor,
        indoor_price,
        outdoor,
        outdoor_price,
        total,
        mxz_model,
        mxz_price,
        d.get("status", "В наличии"),
        item_id
    ))

    conn.commit()
    cur.close()
    conn.close()

    return redirect(admin_url())


@app.route("/admin/delete/<int:item_id>", methods=["POST"])
def delete_item(item_id):
    if not check_admin():
        return "Access denied"

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM catalog WHERE id=%s", (item_id,))
    conn.commit()
    cur.close()
    conn.close()

    return redirect(admin_url())


@app.route("/admin/upload", methods=["POST"])
def upload_excel():
    if not check_admin():
        return "Access denied"

    file = request.files.get("file")
    if not file:
        return redirect(admin_url())

    wb = load_workbook(file)
    ws = wb.active

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM catalog")

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        category = str(row[0] or "").strip().lower()
        series = str(row[1] or "").strip()
        indoor = str(row[2] or "").strip()
        indoor_price = to_int(row[3])
        outdoor = str(row[4] or "").strip()
        outdoor_price = to_int(row[5])
        mxz_model = str(row[6] or "").strip()
        mxz_price = to_int(row[7])
        status = str(row[8] or "В наличии").strip()

        if category == "multisplit":
            series = ""
            indoor = ""
            indoor_price = 0
            outdoor = ""
            outdoor_price = 0
        else:
            category = "split"
            mxz_model = ""
            mxz_price = 0

        total = calc_total(category, indoor_price, outdoor_price, mxz_price)

        cur.execute("""
            INSERT INTO catalog
            (category, series, indoor, indoor_price, outdoor, outdoor_price, total, mxz_model, mxz_price, status)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            category,
            series,
            indoor,
            indoor_price,
            outdoor,
            outdoor_price,
            total,
            mxz_model,
            mxz_price,
            status
        ))

    conn.commit()
    cur.close()
    conn.close()

    return redirect(admin_url())


@app.route("/api/catalog")
def catalog():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM catalog ORDER BY id DESC")
    items = cur.fetchall()
    cur.close()
    conn.close()

    return jsonify(items)


if __name__ == "__main__":
    app.run()
